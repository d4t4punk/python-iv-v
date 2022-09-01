# Scott Newby - 2022-05-06
# Description
# Python & PostreSQL Independent Validation & Verification (IVV)
# Main.py provides an entry point to kick off source & target
# processes and to produce the IVV Excel artifacts
# this project is meant to be a framework for an operator to fill-in with 
# SQL scripts and to be customizable to various sources and targets
# There is a companion postgreSQL to postgreSQL small databawse
# migration project embedded to exercise and become familiar with the framework 
# This project was designed and run under miniconda on Linux


# Imports

import psycopg2
import openpyxl
from openpyxl.styles import Font
from openpyxl import Workbook
import getopt
import sys
import os
import datetime
import yaml
# import other python scripts

import pgconnect

def main(argv):
    print('main')
    # open the config
    with open('config.yml') as cfg:
        config = yaml.safe_load(cfg)

    # Get validation configurations out of the config.yml
    
    ivvdb = config['IVVDB']['db']
    ivvpwd = config['IVVDB']['pwd']
    ivvusr = config['IVVDB']['usr']
    ivvport = config['IVVDB']['port']
    ivvhost = config['IVVDB']['host']
    # Define the global connection
    global pgconnection
    # TODO remove this
    print(ivvdb,ivvpwd,ivvusr,ivvport,ivvhost)
    # Create the connection
    pgconnection = pgconnect.pgconnect(ivvdb,ivvusr,ivvpwd,ivvhost,ivvport)

def export_to_excel(qry, headings, filepath):
    print('export to excel')
    # create the cursor from the global connnection
    local_cursor = pgconnection.cursor()
    # execute the query
    local_cursor.execute(qry)
    # get the data
    exp_data = local_cursor.fetchall()
    # close the cursor
    local_cursor.close()
    # define the workbook
    wb = Workbook()
    # define the sheet
    sheet = wb.active
    # bold the header
    sheet.row_dimensions[1].font = Font(bold = True)
    # write the header
    for col, heading in enumerate(headings, start = 1):
        sheet.cell(row = 1, column = col).value = heading
    # write the data
    for row, rows in enumerate(exp_data, start = 2):
        for col, cell_val in enumerate(rows, start = 1):
            sheet.cell(rows = row, column = col).value = cell_val
    # save the file
    wb.save(filepath)

main(sys.argv[1:])